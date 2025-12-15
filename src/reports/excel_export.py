"""
Excel Export for Geotechnical Analysis Results.
"""

from pathlib import Path
from typing import Dict, Any


def export_to_excel(
    results: Dict,
    soil_params: Dict[str, float],
    foundation_params: Dict[str, float],
    output_path: Path
) -> Path:
    """
    Export analysis results to Excel workbook.
    
    Note: Requires openpyxl library. Falls back to CSV if unavailable.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.chart import LineChart, Reference
        
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        header_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Soil properties
        ws_summary['A1'] = "Soil Properties"
        ws_summary['A1'].font = Font(bold=True, size=14)
        
        row = 3
        for key, value in soil_params.items():
            ws_summary[f'A{row}'] = key.replace('_', ' ').title()
            ws_summary[f'B{row}'] = value
            row += 1
        
        row += 2
        ws_summary[f'A{row}'] = "Foundation Parameters"
        ws_summary[f'A{row}'].font = Font(bold=True, size=14)
        row += 2
        
        for key, value in foundation_params.items():
            ws_summary[f'A{row}'] = key.replace('_', ' ').title()
            ws_summary[f'B{row}'] = value
            row += 1
        
        # Results sheet
        ws_results = wb.create_sheet("Results")
        
        headers = ["Time (days)", "Depth (m)", "σv (kPa)", "σ'v (kPa)", "σ'h (kPa)", "u (kPa)", "FS"]
        for col, header in enumerate(headers, 1):
            cell = ws_results.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        row = 2
        for t in sorted(results.keys()):
            for z in sorted(results[t].keys()):
                d = results[t][z]
                ws_results.cell(row=row, column=1, value=t)
                ws_results.cell(row=row, column=2, value=z)
                ws_results.cell(row=row, column=3, value=d.get('sigma_v', 0))
                ws_results.cell(row=row, column=4, value=d.get('sigma_v_eff', 0))
                ws_results.cell(row=row, column=5, value=d.get('sigma_h_eff', 0))
                ws_results.cell(row=row, column=6, value=d.get('pore_pressure', 0))
                ws_results.cell(row=row, column=7, value=d.get('fs', 0))
                row += 1
        
        # Adjust column widths
        for ws in [ws_summary, ws_results]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = max_length + 2
        
        wb.save(str(output_path))
        return output_path
        
    except ImportError:
        # Fallback to CSV
        csv_path = output_path.with_suffix('.csv')
        with open(csv_path, 'w') as f:
            f.write("time_days,depth_m,sigma_v,sigma_v_eff,sigma_h_eff,pore_pressure,fs\n")
            for t in sorted(results.keys()):
                for z in sorted(results[t].keys()):
                    d = results[t][z]
                    f.write(f"{t},{z},{d.get('sigma_v',0):.2f},{d.get('sigma_v_eff',0):.2f},")
                    f.write(f"{d.get('sigma_h_eff',0):.2f},{d.get('pore_pressure',0):.2f},{d.get('fs',0):.4f}\n")
        return csv_path
